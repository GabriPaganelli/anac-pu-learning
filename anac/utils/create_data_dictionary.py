"""
create_data_dictionary.py
Genera variables/data_dictionary.xlsx con tutte le colonne di bando_cig_all.parquet.

Colonne del foglio:
  variabile, descrizione, dataset_fonte, fase_modello, tipo_dato,
  null_pct_full, null_pct_labeled, trattamento_na, note_segnale
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

_BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(_BASE, "variables", "data_dictionary.xlsx")

# Dati null% calcolati da bando_cig_all.parquet (9,468,795 righe, 2,598 labeled)
# formato: variabile -> (null%_full, null%_labeled)
NULL_PCT = {
    "cig": (0.0, 0.0),
    "label": (100.0, 0.0),
    "esito": (60.3, 39.0),
    "anno_pubblicazione": (0.0, 0.0),
    "regione": (35.0, 25.4),
    "importo_complessivo_gara": (0.1, 0.0),
    "n_lotti_componenti": (14.0, 1.7),
    "importo_lotto": (0.0, 0.0),
    "oggetto_principale_contratto": (0.0, 0.0),
    "cod_modalita_realizzazione": (0.0, 0.0),
    "sezione_regionale": (1.5, 2.8),
    "cod_strumento_svolgimento": (76.8, 49.2),
    "flag_urgenza": (0.0, 0.0),
    "cod_motivo_urgenza": (89.5, 65.3),
    "flag_delega": (91.7, 56.5),
    "finestra_offerta_giorni": (0.3, 0.4),
    "lag_perfezionamento_giorni": (0.0, 0.0),
    "lag_comunicazione_esito_giorni": (60.3, 40.9),
    "flag_accordo_quadro": (0.0, 0.0),
    "flag_ripetizioni": (0.0, 0.0),
    "settore_speciale": (0.0, 0.0),
    "flag_appalto_riservato": (0.0, 0.0),
    "importo_sicurezza_pct": (73.9, 35.6),
    "tipo_scelta_4cls": (0.0, 0.6),
    "cpv_macro_categoria": (0.0, 0.0),
    "tasso_disoccupazione": (35.0, 25.4),
    "reddito_irpef_procapite": (35.0, 25.4),
    "tasso_omicidi_100k": (35.0, 25.4),
    "numero_offerte_ammesse": (73.7, 39.6),
    "pct_offerte_escluse": (83.2, 64.4),
    "ribasso_aggiudicazione": (62.8, 41.1),
    "ribasso_spread": (88.8, 78.3),
    "flag_vince_minimo": (62.5, 38.0),
    "flag_progettazione_esterna": (62.5, 38.0),
    "lag_aggiudicazione_giorni": (70.4, 46.1),
    "asta_elettronica": (0.0, 0.0),
    "flag_scomputo": (0.0, 0.0),
    "flag_proc_accelerata": (0.0, 0.0),
    "importo_aggiudicazione": (62.6, 39.1),
    "num_imprese_offerenti": (73.7, 39.7),
    "numero_offerte_escluse": (74.3, 62.2),
    "tipo_soggetto_agg": (62.6, 40.5),
    "natura_giuridica_SA": (0.1, 0.0),
    "lag_stipula_aggiudicazione_giorni": (89.8, 83.3),
    "durata_pianificata_giorni": (87.3, 80.5),
    "consegna_frazionata": (84.4, 79.1),
    "consegna_sotto_riserva": (84.4, 79.1),
    "n_varianti": (0.0, 0.0),
    "flag_variante_sostanziale": (0.0, 0.0),
    "pct_overrun_variante": (97.8, 95.8),
    "pct_vita_prima_variante": (98.3, 96.5),
    "flag_variante_oltre_termine": (0.0, 0.0),
    "pct_riserva_base": (72.0, 45.6),
    "pct_overrun_core": (93.6, 96.7),
    "pct_riserva_consumata": (93.6, 96.7),
    "n_sospensioni": (0.0, 0.0),
    "durata_totale_sospensioni_gg": (0.0, 0.0),
    "flag_sosp_giudiziaria": (0.0, 0.0),
    "flag_sospensione": (0.0, 0.0),
    "pct_durata_sospesa": (0.2, 0.2),
    "n_sal": (0.0, 0.0),
    "flag_in_ritardo": (96.8, 92.3),
    "flag_proroga": (0.0, 0.0),
    "flag_subappalto": (0.0, 0.0),
    "tipo_lavorazione_macro": (90.9, 81.6),
    "esito_collaudo": (93.5, 96.6),
}

# Dizionario principale
# Ogni voce: (descrizione, dataset_fonte, fase_modello, tipo_dato, trattamento_na, note_segnale)
# fase_modello: METADATA | M1 | M2 | M3 | M3_SOLO_NATIVI
COLS = [
    ("cig",
     "Codice identificativo gara (chiave primaria)",
     "BANDO CIG", "METADATA", "string",
     "—",
     "Identificatore univoco; sempre presente"),

    ("label",
     "Label PU learning: 1=condannato, 0=scagionato, NaN=unlabeled",
     "LABELS (OpenGA + ANAC fine_contratto)", "METADATA", "int64",
     "—",
     "781 positivi, 1817 negativi, ~9.47M unlabeled"),

    ("esito",
     "Esito della procedura di gara (AGGIUDICATA / ANNULLATA / None / ...)",
     "BANDO CIG", "METADATA", "string",
     "—",
     "Colonna di selezione righe per M2/M3; non usata come feature"),

    ("anno_pubblicazione",
     "Anno di pubblicazione del bando",
     "BANDO CIG", "METADATA", "int64",
     "—",
     "Metadata per analisi temporali; non feature di modello"),

    ("regione",
     "Regione italiana dell'appalto (derivata da luogo_istat)",
     "BANDO CIG + ISTAT", "METADATA", "string",
     "NaN se luogo_istat assente (~35%)",
     "Metadata per analisi geografiche; non feature di modello"),

    ("importo_complessivo_gara",
     "Valore totale della gara (somma di tutti i lotti)",
     "BANDO CIG", "M1", "float64",
     "Nativi: NA nativo; non-nativi: discretizzazione log-quantile",
     "Feature base; invariante per gare mono-lotto"),

    ("n_lotti_componenti",
     "Numero di lotti della gara (14% null → imputato a 1)",
     "BANDO CIG", "M1", "int64",
     "14% null; nativi: imputa 1; non-nativi: idem",
     "Proxy complessità procedurale"),

    ("importo_lotto",
     "Valore economico del lotto (base d'asta)",
     "BANDO CIG", "M1", "float64",
     "Nativi: NA nativo (<0.1%); non-nativi: discretizzazione log-quantile",
     "Feature principale di importo"),

    ("oggetto_principale_contratto",
     "Tipo contratto: LAVORI / SERVIZI / FORNITURE",
     "BANDO CIG", "M1", "string",
     "0% null; sempre presente",
     "Categoriale 3 valori"),

    ("cod_modalita_realizzazione",
     "Modalità di realizzazione (codice ANAC)",
     "BANDO CIG", "M1", "int64",
     "0% null; categoriale",
     "Alta cardinalità; considerare riduzione"),

    ("sezione_regionale",
     "Sezione ANAC: CENTRALE / REGIONALE / NaN",
     "BANDO CIG", "M1", "string",
     "1.5% null; trattare come categoria MISSING",
     "—"),

    ("cod_strumento_svolgimento",
     "Strumento di svolgimento della gara (76.8% null → categoria MISSING)",
     "BANDO CIG", "M1", "int64",
     "76.8% null; nativi: NA nativo; non-nativi: categoria MISSING",
     "—"),

    ("flag_urgenza",
     "Presenza di procedura d'urgenza (0/1)",
     "BANDO CIG", "M1", "int64",
     "0% null",
     "—"),

    ("cod_motivo_urgenza",
     "Codice motivo urgenza (89.5% null → categoria MISSING)",
     "BANDO CIG", "M1", "int64",
     "89.5% null; nativi: NA nativo; non-nativi: categoria MISSING",
     "—"),

    ("flag_delega",
     "Contratto delegato a centrale di committenza (91.7% null → 3ª categoria)",
     "BANDO CIG", "M1", "int64",
     "91.7% null; trattare come 3ª categoria (0/1/MISSING)",
     "—"),

    ("finestra_offerta_giorni",
     "Giorni dalla pubblicazione alla scadenza offerte",
     "BANDO CIG", "M1", "float64",
     "0.3% null; nativi: NA nativo; non-nativi: discretizzazione",
     "Finestre molto brevi → possibile restrizione accesso"),

    ("lag_perfezionamento_giorni",
     "Giorni dalla pubblicazione al perfezionamento del bando",
     "BANDO CIG", "M1", "float64",
     "0% null",
     "—"),

    ("flag_accordo_quadro",
     "Flag accordo quadro (0/1)",
     "BANDO CIG", "M1", "int8",
     "0% null",
     "—"),

    ("flag_ripetizioni",
     "Flag ripetizioni previste (0/1)",
     "BANDO CIG", "M1", "int8",
     "0% null",
     "—"),

    ("settore_speciale",
     "Settore speciale ex D.Lgs. 50/2016 (0/1)",
     "BANDO CIG", "M1", "int8",
     "0% null",
     "—"),

    ("flag_appalto_riservato",
     "Appalto riservato a categorie protette (0/1)",
     "BANDO CIG", "M1", "int8",
     "0% null",
     "—"),

    ("importo_sicurezza_pct",
     "Quota oneri sicurezza / importo lotto (73.9% null; fallback da QE BASE_ASTA)",
     "BANDO CIG + QE BASE_ASTA", "M1", "float64",
     "73.9% null; nativi: NA nativo; non-nativi: mediana-impute",
     "COND mediana 0.11 vs SCAG 0.05 (3.79x) — segnale forte"),

    ("tipo_scelta_4cls",
     "Procedura di scelta contraente: APERTA / RISTRETTA / NEGOZIATA / AFFIDAMENTO_DIRETTO",
     "BANDO CIG + lookup", "M1", "string",
     "0% null; 0.6% null su labeled",
     "Mapping 35+ codici → 4 classi; AFFIDAMENTO_DIRETTO ~55% del totale"),

    ("cpv_macro_categoria",
     "Macro-categoria CPV: LAVORI / SERVIZI / FORNITURE / IT / ING_PROF / SANITA",
     "BANDO CIG + CPV lookup", "M1", "string",
     "0% null",
     "Prime 2 cifre CPV → 6 macro-categorie"),

    ("tasso_disoccupazione",
     "Tasso di disoccupazione provinciale (anno t-1)",
     "ISTAT (t-1)", "M1", "float64",
     "35% null (luogo_istat assente); nativi: NA nativo; non-nativi: mediana provinciale",
     "Proxy contesto economico locale"),

    ("reddito_irpef_procapite",
     "Reddito IRPEF pro-capite provinciale (anno t-1, €)",
     "MEF (t-1)", "M1", "float64",
     "35% null; nativi: NA nativo; non-nativi: mediana provinciale",
     "Proxy ricchezza locale"),

    ("tasso_omicidi_100k",
     "Tasso omicidi per 100k abitanti, provinciale",
     "ISTAT BES", "M1", "float64",
     "35% null; nativi: NA nativo; non-nativi: mediana provinciale",
     "Proxy criminalità locale; backfill max 10 anni"),

    ("pct_riserva_base",
     "Somme a disposizione SA / costi core (QE BASE_ASTA) — proxy margine contingency",
     "QUADRO ECONOMICO (BASE_ASTA)", "M1", "float64",
     "72% null; nativi: NA nativo; non-nativi: mediana-impute",
     "COND mediana 0.22 vs SCAG 0.09 (2.6x) — segnale forte ante"),

    ("tipo_lavorazione_macro",
     "Tipo lavorazione: COSTRUZIONE / RISANAMENTO / MANUTENZIONE (solo LAVORI; 90.9% MISSING)",
     "LAVORAZIONI + lookup", "M1", "string",
     "90.9% null → MISSING per non-LAVORI; trattare come categoria",
     "MANUTENZIONE ratio 2.2x; COSTRUZIONE 1.9x"),

    ("natura_giuridica_SA",
     "Natura giuridica stazione appaltante: PA_ENTE / SOCIETA / PA_ISTITUZIONALE / SANITARIO / EPE / NON_PROFIT / CONSORZIO / ALTRO",
     "STAZIONE APPALTANTE", "M1", "string",
     "0.1% null; quasi completo",
     "SANITARIO: solo 355 SA ma 10% CIG — alta esposizione corruttiva"),

    ("numero_offerte_ammesse",
     "Numero di offerte ammesse alla gara",
     "AGGIUDICAZIONI", "M2", "float64",
     "73.7% full (0% in M2); nativi: NA nativo; non-nativi: mediana-impute",
     "↑ COND: cartello — offerte fasulle per simulare competizione"),

    ("numero_offerte_escluse",
     "Numero di offerte escluse",
     "AGGIUDICAZIONI", "M2", "float64",
     "74.3% full (bassa in M2); nativi: NA nativo",
     "—"),

    ("num_imprese_offerenti",
     "Numero di imprese che hanno presentato offerta",
     "AGGIUDICAZIONI", "M2", "float64",
     "73.7% full (0% in M2); nativi: NA nativo",
     "Fonte autorevole (più completa di partecipanti.csv)"),

    ("pct_offerte_escluse",
     "Percentuale offerte escluse (escluse / max(offerenti, ammesse+escluse))",
     "AGGIUDICAZIONI", "M2", "float64",
     "83.2% full; nativi: NA nativo",
     "—"),

    ("ribasso_aggiudicazione",
     "Ribasso percentuale offerto dal vincitore",
     "AGGIUDICAZIONI", "M2", "float64",
     "62.8% full (bassa in M2); nativi: NA nativo",
     "—"),

    ("ribasso_spread",
     "Differenza tra ribasso massimo e ribasso aggiudicatario (proxy competition)",
     "AGGIUDICAZIONI", "M2", "float64",
     "88.8% full; NaN se non asta al ribasso; nativi: NA nativo",
     "—"),

    ("flag_vince_minimo",
     "Il vincitore ha offerto il ribasso minimo? (0/1; NaN se non comparabile)",
     "AGGIUDICAZIONI", "M2", "bool",
     "62.5% full; trattare NaN come MISSING",
     "—"),

    ("flag_progettazione_esterna",
     "Progettazione affidata esternamente? (0/1; NaN se non applicabile)",
     "AGGIUDICAZIONI", "M2", "bool",
     "62.5% full; nativi: NA nativo; non-nativi: trattare NaN come MISSING",
     "—"),

    ("lag_aggiudicazione_giorni",
     "Giorni dalla pubblicazione all'aggiudicazione definitiva",
     "AGGIUDICAZIONI + BANDO CIG", "M2", "float64",
     "70.4% full; NaN se negativo (caricamento retroattivo ANAC); nativi: NA nativo",
     "SC-02: 18.5% lag negativi → NaN"),

    ("lag_comunicazione_esito_giorni",
     "Giorni dalla pubblicazione alla comunicazione esito gara",
     "BANDO CIG", "M2", "float64",
     "60.3% full (= % esito null); nativi: NA nativo",
     "—"),

    ("asta_elettronica",
     "Procedura condotta su piattaforma di asta elettronica (0/1)",
     "AGGIUDICAZIONI", "M2", "int8",
     "0% null (NaN → 0: assenza di aggiudicazioni = no asta elettronica)",
     "—"),

    ("flag_scomputo",
     "Scomputo oneri sicurezza (0/1)",
     "AGGIUDICAZIONI", "M2", "int8",
     "0% null (NaN → 0)",
     "—"),

    ("flag_proc_accelerata",
     "Procedura accelerata ex art. 163 D.Lgs. 50/2016 (0/1)",
     "AGGIUDICAZIONI", "M2", "int8",
     "0% null (NaN → 0)",
     "—"),

    ("importo_aggiudicazione",
     "Importo definitivo di aggiudicazione (€)",
     "AGGIUDICAZIONI", "M2", "float64",
     "62.6% full (0% in M2); nativi: NA nativo",
     "—"),

    ("tipo_soggetto_agg",
     "Tipo soggetto aggiudicatario: SINGOLA / ATI / CONSORZIO / SA / GEIE / ALTRO",
     "AGGIUDICATARI", "M2", "string",
     "62.6% full (0% in M2); trattare NaN come MISSING",
     "—"),

    ("flag_subappalto",
     "Presenza di subappalto (OR tra flag bando e registrazione in subappalti.csv)",
     "AGGIUDICAZIONI + SUBAPPALTI", "M2", "int64",
     "0% null (NaN → 0)",
     "COND 21.0% vs SCAG 17.0% (1.2x) — moderato"),

    ("lag_stipula_aggiudicazione_giorni",
     "Giorni tra aggiudicazione e firma del contratto",
     "AVVIO CONTRATTO + AGGIUDICAZIONI", "M3", "float64",
     "89.8% full; NaN se negativo (caricamento retroattivo); nativi: NA nativo",
     "—"),

    ("durata_pianificata_giorni",
     "Durata prevista del contratto in giorni (termine - inizio effettiva)",
     "AVVIO CONTRATTO", "M3", "float64",
     "87.3% full; nativi: NA nativo",
     "Denominatore per pct_durata_sospesa e pct_vita_prima_variante"),

    ("consegna_frazionata",
     "Consegna frazionata in più tranche (0/1/NaN come 3ª categoria)",
     "AVVIO CONTRATTO", "M3", "int8",
     "84.4% full; NaN = dato non registrato; trattare come 3ª categoria",
     "—"),

    ("consegna_sotto_riserva",
     "Consegna sotto riserva (0/1/NaN come 3ª categoria)",
     "AVVIO CONTRATTO", "M3", "int8",
     "84.4% full; NaN = dato non registrato; trattare come 3ª categoria",
     "—"),

    ("n_varianti",
     "Numero di varianti approvate (0 se assente, cap 20)",
     "VARIANTI", "M3", "int64",
     "0% null (assenza = 0, non NaN)",
     "mediana=1, p95=3 per CIG con varianti"),

    ("flag_variante_sostanziale",
     "Almeno una variante sostanziale (codici non amministrativi) (0/1)",
     "VARIANTI", "M3", "int8",
     "0% null (assenza = 0)",
     "COND 4.7% vs SCAG 0.9% (5.2x) — segnale forte"),

    ("pct_overrun_variante",
     "Importo varianti (da QE) / importo aggiudicazione (97.8% null)",
     "QUADRO ECONOMICO (VARIANTE)", "M3_SOLO_NATIVI", "float64",
     "97.8% null; SOLO modelli nativi (XGBoost/LightGBM)",
     "COND mediana 1.78 vs SCAG 1.45; analogo binario: flag_variante_sostanziale"),

    ("pct_vita_prima_variante",
     "Giorni prima variante / durata pianificata — quando è arrivata la prima variante",
     "VARIANTI + AVVIO CONTRATTO", "M3_SOLO_NATIVI", "float64",
     "98.3% null; SOLO modelli nativi",
     "Analogo binario: flag_variante_oltre_termine"),

    ("flag_variante_oltre_termine",
     "Prima variante arrivata dopo la scadenza contrattuale (0/1)",
     "VARIANTI + AVVIO CONTRATTO", "M3", "int64",
     "0% null (assenza varianti = 0)",
     "COND ~4.6% vs SCAG ~1.0%"),

    ("pct_overrun_core",
     "Costo finale (QE CONSUNTIVO) / costo iniziale (QE BASE_ASTA) — cost overrun",
     "QUADRO ECONOMICO", "M3", "float64",
     "93.6% full; valori <0 o >10 → NaN; nativi: NA nativo",
     "Sample piccolo (73 COND / 14 SCAG); interpretazione con cautela"),

    ("pct_riserva_consumata",
     "Somme a disposizione consumate (CONSUNTIVO) / costi core iniziali (BASE_ASTA)",
     "QUADRO ECONOMICO", "M3", "float64",
     "93.6% full; valori <0 o >10 → NaN; nativi: NA nativo",
     "Sample piccolo; possibile bias da appalti interrotti"),

    ("n_sospensioni",
     "Numero di sospensioni lavori (0 se assente)",
     "SOSPENSIONI", "M3", "int64",
     "0% null (assenza = 0)",
     "—"),

    ("durata_totale_sospensioni_gg",
     "Durata totale delle sospensioni in giorni (0 se assente; escluse sospensioni aperte)",
     "SOSPENSIONI", "M3", "float64",
     "0% null (assenza = 0)",
     "—"),

    ("flag_sosp_giudiziaria",
     "Almeno una sospensione per ordine giudiziario (0/1)",
     "SOSPENSIONI", "M3", "int64",
     "0% null",
     "534 CIG; unico motivo con implicazione legale diretta esplicita"),

    ("flag_sospensione",
     "Almeno una sospensione (0/1)",
     "SOSPENSIONI", "M3", "int64",
     "0% null",
     "COND 5.9% vs SCAG 0.7% (8.4x) — segnale molto forte"),

    ("pct_durata_sospesa",
     "Durata totale sospensioni / durata pianificata (cap 10; 0 se no sospensioni)",
     "SOSPENSIONI + AVVIO CONTRATTO", "M3", "float64",
     "0.2% null (solo quando sospensioni presenti ma durata mancante)",
     "pct > 1 atteso (sospensioni estendono contratto)"),

    ("n_sal",
     "Numero di stati di avanzamento lavori registrati (0 se assente)",
     "SAL", "M3", "int64",
     "0% null (assenza = 0)",
     "Coverage 3.2% parquet; LAVORI 11.4%, SERVIZI 2.4%"),

    ("flag_in_ritardo",
     "Almeno un SAL in ritardo (0/1/NaN se nessun SAL — 3ª categoria)",
     "SAL", "M3", "int8",
     "96.8% full; NaN = nessun SAL registrato; trattare come 3ª categoria",
     "COND 14.7% vs SCAG 7.7% (1.9x; calcolato su CIG con SAL)"),

    ("flag_proroga",
     "Almeno una proroga registrata nei SAL (0/1)",
     "SAL", "M3", "int64",
     "0% null",
     "COND 1.7% vs SCAG 0.2% (8.5x) — segnale molto forte"),

    ("tipo_lavorazione_macro",  # già a M1
     "", "", "", "", "", ""),   # duplicato — gestito sotto

    ("esito_collaudo",
     "Esito del collaudo: POSITIVO / NEGATIVO / NaN (93.5% null = non registrato)",
     "COLLAUDO", "M3", "string",
     "93.5% null; NaN ≠ negativo; trattare come categoria con MISSING",
     "POSITIVO ratio 2.73x: selection bias (condannati completano iter formale)"),
]

# Rimuovi il placeholder duplicato tipo_lavorazione_macro inserito per errore
COLS = [c for c in COLS if not (c[0] == "tipo_lavorazione_macro" and c[1] == "")]

HEADERS = [
    "variabile", "descrizione", "dataset_fonte", "fase_modello",
    "tipo_dato", "null%_full", "null%_labeled",
    "trattamento_na", "note_segnale"
]

# Colori per fase modello
FASE_COLORS = {
    "METADATA":       "D9E1F2",  # blu chiaro
    "M1":             "E2EFDA",  # verde chiaro
    "M2":             "FFF2CC",  # giallo chiaro
    "M3":             "FCE4D6",  # arancione chiaro
    "M3_SOLO_NATIVI": "F4CCCC",  # rosso chiaro
}

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "data_dictionary"

# Header row
header_fill = PatternFill("solid", fgColor="2F5496")
header_font = Font(bold=True, color="FFFFFF", size=11)
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

for col_idx, h in enumerate(HEADERS, 1):
    cell = ws.cell(row=1, column=col_idx, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_align

ws.row_dimensions[1].height = 30

# Data rows
for row_idx, entry in enumerate(COLS, 2):
    varname, descr, fonte, fase, dtype, na_treat, note = entry
    null_full, null_lbl = NULL_PCT.get(varname, ("—", "—"))

    values = [varname, descr, fonte, fase, dtype,
              null_full, null_lbl, na_treat, note]

    bg_color = FASE_COLORS.get(fase, "FFFFFF")
    fill = PatternFill("solid", fgColor=bg_color)

    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.fill = fill
        cell.alignment = Alignment(vertical="top", wrap_text=(col_idx in (2, 8, 9)))

    ws.row_dimensions[row_idx].height = 40

# Column widths
COL_WIDTHS = [28, 52, 28, 18, 12, 13, 15, 48, 52]
for i, w in enumerate(COL_WIDTHS, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Freeze top row
ws.freeze_panes = "A2"

# Add legend sheet
ls = wb.create_sheet("legenda")
legend = [
    ("METADATA", "Colonne di metadata/identificazione — escluse da tutti i feature set"),
    ("M1",       "Feature ex ante — disponibili alla pubblicazione del bando"),
    ("M2",       "Feature durante — disponibili dopo l'aggiudicazione (M2 = M1 ∪ M2_new)"),
    ("M3",       "Feature ex post — disponibili dopo stipula/esecuzione (M3 = M2 ∪ M3_new)"),
    ("M3_SOLO_NATIVI", "Feature M3 usate solo da XGBoost/LightGBM (alta null%; hanno analogo binario)"),
]
ls.cell(1, 1, "fase_modello").font = Font(bold=True)
ls.cell(1, 2, "significato").font = Font(bold=True)
for i, (fase, desc) in enumerate(legend, 2):
    c = ls.cell(i, 1, fase)
    c.fill = PatternFill("solid", fgColor=FASE_COLORS[fase])
    ls.cell(i, 2, desc)
ls.column_dimensions["A"].width = 20
ls.column_dimensions["B"].width = 70

wb.save(OUT)
print(f"Salvato: {OUT}")
print(f"Righe dati: {len(COLS)}")
